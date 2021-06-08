import os, sys

# 获取当前路径， 通过anchor文件获取项目root路径
this_file_path = os.path.split(os.path.realpath(__file__))[0]
this_path = this_file_path
while this_path:
    if os.path.exists(os.path.join(this_path, 'sosweety_root_anchor.py')):
        root_path = this_path
        break
    par_path = os.path.dirname(this_path)
    # print(par_path)
    if par_path == this_path:
        root_path = this_file_path
        break
    else:
        this_path = par_path
sys.path.append(root_path)

import openpyxl
# import pandas as pd
import json
import csv
from modules.knowledgebase.kb import KnowledgeBase
from modules.grammarbase.grammerbase_class import PhrasePattern, SubSentencePattern
from modules.grammarbase.grammarbase import GrammarBase


class Updater(object):
    def __init__(self, KB):
        self.KB = KB
        return

    # 如果不想重新做kb文件，可以先单独备份下kb
    def update(self, train_dir):
        GB = GrammarBase()
        phrase_dirs = ['hf_word_phrase', 'noun_phrase']
        ss_dirs = ['ss_pattern']
        analyse_result_dir = os.path.join(train_dir, 'analyse_result')

        def merge_xlsx_file(old_file_path, new_file_path, pattern_type):
            # backup old file
            old_bak_file_path = old_file_path + '.bak'
            os.popen('cp ' + old_file_path + ' ' + old_bak_file_path)

            # old file 用csv追加写入
            old_file = open(old_file_path, 'a+')
            old_file_writer = csv.writer(old_file)

            # 构建pattern 唯一标识符
            if pattern_type == 'phrase':
                phrase_pattern_str_set = set()
                for phrase_pattern in GB.phrase_patterns:
                    phrase_pattern_str = json.dumps(phrase_pattern.features, ensure_ascii=False)
                    phrase_pattern_str_set.add(phrase_pattern_str)

            if pattern_type == 'ss':
                ss_pattern_str_set = set()
                for ss_pattern in GB.ss_patterns:
                    ss_pattern_str = json.dumps(ss_pattern.parse_str, ensure_ascii=False) + '_' + ss_pattern.ss_type
                    ss_pattern_str_set.add(ss_pattern_str)

            # new file 用xls读入
            new_xl_file = openpyxl.load_workbook(new_file_path)
            sheetnames = new_xl_file.get_sheet_names()
            ws = new_xl_file[sheetnames[0]]
            for row in ws.iter_rows(min_row=2):
                value_list = [cell.value for cell in row]
                try:
                    if pattern_type == 'phrase':
                        tmp_pattern = PhrasePattern(value_list[0], value_list[1], value_list[2], value_list[3], value_list[4], value_list[5], value_list[6])
                        tmp_pattern_str = json.dumps(tmp_pattern.features, ensure_ascii=False)
                        if tmp_pattern_str not in phrase_pattern_str_set:
                            old_file_writer.writerow(value_list)
                    if pattern_type == 'ss':
                        tmp_pattern = SubSentencePattern(value_list[0], value_list[1], value_list[2], value_list[3], value_list[4], value_list[5])
                        tmp_pattern_str = json.dumps(tmp_pattern.parse_str, ensure_ascii=False) + '_' + tmp_pattern.ss_type
                        if tmp_pattern_str not in ss_pattern_str_set:
                            old_file_writer.writerow(value_list)
                except(Exception):
                    print(value_list)
            return

        # phrase patterns
        new_concep_relation_set = set()
        for phrase_dir in phrase_dirs:
            phrase_dir = os.path.join(analyse_result_dir, phrase_dir)

            # new concept
            # todo frequency怎么加？
            new_concept_file = os.path.join(phrase_dir, 'identified_new_concept.xlsx')
            if os.path.exists(new_concept_file):
                xl_file = openpyxl.load_workbook(new_concept_file)
                sheetnames = xl_file.get_sheet_names()
                ws = xl_file.get_sheet_by_name(sheetnames[0])
                for row in ws.iter_rows():
                    word = row[1].value
                    if word is not None and word != '':
                        new_concep_relation_set.add(json.dumps([word], ensure_ascii=False))
                        if len(row) > 4 and row[4].value is not None and row[4].value != '':
                            # print(row[4].value)
                            new_concep_relation_set.add(json.dumps([word, row[4].value], ensure_ascii=False))

            # new upper concept
            new_upper_concept_file = os.path.join(phrase_dir, 'identified_new_upper_concept.xlsx')
            if os.path.exists(new_upper_concept_file):
                xl_file = openpyxl.load_workbook(new_upper_concept_file)
                sheetnames = xl_file.get_sheet_names()
                ws = xl_file.get_sheet_by_name(sheetnames[0])
                for row in ws.iter_rows():
                    concept_list = row[0].value.split('|')
                    if len(row) > 3:
                        for concept in concept_list:
                            new_concep_relation_set.add(json.dumps([concept, row[3].value], ensure_ascii=False))
                    else:
                        for concept in concept_list:
                            new_concep_relation_set.add(json.dumps([concept, row[0].value], ensure_ascii=False))

            # new phrase patterns
            new_phrase_pattern_file = os.path.join(phrase_dir, 'identified_new_phrase_pattern.xlsx')
            if os.path.exists(new_phrase_pattern_file):
                phrase_pattern_file_name = 'modules/grammarbase/base_files/phrase_pattern.csv'
                phrase_pattern_file_path = os.path.join(root_path, phrase_pattern_file_name)
                merge_xlsx_file(phrase_pattern_file_path, new_phrase_pattern_file, 'phrase')

        # add relations to kb and files
        kb_file_name = 'data/init_data/kb_relations/added_relations'
        kb_bak_file_name = 'data/init_data/kb_relations/added_relations.bak'
        kb_file_path = os.path.join(root_path, kb_file_name)
        kb_bak_file_path = os.path.join(root_path, kb_bak_file_name)
        if os.path.exists(kb_file_path):
            os.popen('cp ' + kb_file_path + ' ' + kb_bak_file_path)
        kb_file = open(kb_file_path, 'a+')
        for relation in new_concep_relation_set:
            relation = json.loads(relation)
            self.KB.add_relation(relation)
            if len(relation) > 1:
                kb_file.write(relation[0] + '\t' + relation[1] + '\n')
            else:
                kb_file.write(relation[0] + '\n')
        kb_file.close()

        # subsentence patterns
        for ss_dir in ss_dirs:
            ss_dir = os.path.join(analyse_result_dir, ss_dir)

            # 目前只有一个dir
            new_ss_pattern_file = os.path.join(ss_dir, 'identified_new_ss_pattern.xlsx')
            if os.path.exists(new_ss_pattern_file):
                ss_pattern_file_name = 'modules/grammarbase/base_files/ss_pattern.csv'
                ss_pattern_file_path = os.path.join(root_path, ss_pattern_file_name)
                merge_xlsx_file(ss_pattern_file_path, new_ss_pattern_file, 'ss')

        # GB = GrammarBase()
        # GB.update(gb_update_dict)
        return


# test
if __name__ == '__main__':
    train_dir = 'data/train_question'
    train_dir = os.path.join(root_path, train_dir)
    KB = KnowledgeBase()
    updater = Updater(KB)
    updater.update(train_dir)
